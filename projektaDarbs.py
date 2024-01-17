import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import re
from openpyxl import Workbook, load_workbook
print("Ievadiet velamas lietas nosaukumu un kvalitati (piem. Butterfly Knife Autotronic Field Tested):")
item = input()

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

urlExchangeRate = "https://www.ecb.europa.eu/stats/policy_and_exchange_rates/euro_reference_exchange_rates/html/index.en.html"
driver.get(urlExchangeRate)

exchangeRateFind = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="main-wrapper"]/main/div[3]/div[2]/div/div/table/tbody/tr[1]/td[3]/a/span')))
exchageRateGet = exchangeRateFind.get_attribute('innerText')
exchangeRate = float(exchageRateGet)

DMarket=[]

urlDM = "https://dmarket.com/ingame-items/item-list/csgo-skins"
driver.get(urlDM)

searchDM = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "o-filter__searchInput")))
searchDM.clear()
searchDM.send_keys(item)

time.sleep(2)
priceFindDM = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "c-asset__priceNumber")))
priceGetDM = priceFindDM.get_attribute('innerText')
priceDM = float(re.sub('[^0-9.]', '', priceGetDM))

commissionDm = "3.85% + €0.27"

withoutCommissionDM = round((priceDM/exchangeRate), 2)
withCommissionDM = round(((withoutCommissionDM)*1.0385 + 0.27), 2)

minDM = 1/exchangeRate
maxDM = 500/exchangeRate

if minDM<=withoutCommissionDM<=maxDM:
    limitsDM = "In"
else:
    limitsDM = "Out"

DMarket.extend(["DMarket: ", withoutCommissionDM, commissionDm, withCommissionDM, limitsDM])

#print(DMarket)

GamerPay = []

urlGP = "https://gamerpay.gg/"
driver.get(urlGP)

acceptCookiesGP= WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID,"CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll")))
acceptCookiesGP.click()

searchGP = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "searchBar")))
searchGP.send_keys(item)
searchGP.send_keys(Keys.ENTER)

time.sleep(2)
priceFindGP = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "ItemCardNewBody_pricePrimary__Tpkw7")))
priceGetGP = priceFindGP.get_attribute('innerText')
priceGP = float(re.sub('[^0-9.]', '', priceGetGP))

commissionGP = "5% + €0.70"

withoutCommissionGP = priceGP
withCommissionGP = round(((withoutCommissionGP)*1.05 + 0.7),2)

limitsGP = "In"

GamerPay.extend(["GamerPay: ", withoutCommissionGP, commissionGP, withCommissionGP, limitsGP])

#print(GamerPay)

SkinBaron = []

urlSB = "https://skinbaron.de/en"
driver.get(urlSB)

acceptCookiesSB = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID,"onetrust-accept-btn-handler")))
acceptCookiesSB.click()

searchSB = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[pathtodata="variants"]')))
searchSB.send_keys(item)
searchSB.send_keys(Keys.ENTER)

time.sleep(2)
priceFindSB = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="offer-container"]/ul/li[1]/sb-offer-card/div/div/div[3]/div/div/span[3]')))
priceGetSB = priceFindSB.get_attribute('innerText')
priceSB = float(re.sub('[^0-9.]', '', priceGetSB))

commissionSB = "5% + €0.35"

withoutCommissionSB = priceSB
withCommissionSB = round(((withoutCommissionSB)*1.05 + 0.35),2)

minSB = 1
maxSB = 7500

if 1<=withoutCommissionSB<=7500:
    limitsSB = "In"
else:
    limitsSB = "Out"

SkinBaron.extend(["SkinBaron: ", withoutCommissionSB, commissionSB, withCommissionSB, limitsSB])

#print(SkinBaron)

title = []
title.extend(["Market Place", "Price Without Commission (€/Eur)", "Commission", "Price With Commission (€/Eur)", "Price In/Out of Deposit Range"])
wb = Workbook()
result = wb.active

result.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(title))
result.cell(row=1, column=1, value=item)

result.append(title)
result.append(DMarket)
result.append(GamerPay)
result.append(SkinBaron)

wb.save("result.xlsx")
wb.close()
driver.quit()